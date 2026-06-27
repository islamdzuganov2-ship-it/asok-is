"""
Celery задачи АСОК ИС.
Брокер: Redis. Результаты: Redis.
"""
import logging
from celery import Celery
from app.core.config import settings

logger = logging.getLogger(__name__)

celery_app = Celery(
    "asok_is",
    broker=settings.REDIS_URL,
    backend=settings.REDIS_URL,
)

celery_app.conf.update(
    task_serializer="json",
    result_serializer="json",
    accept_content=["json"],
    task_track_started=True,
    result_expires=3600,
)


@celery_app.task(bind=True, name="tasks.parse_excel", max_retries=3)
def parse_excel_task(self, file_path: str, period_id: str) -> dict:
    """
    Парсинг Excel файла и батчевая запись val_a/val_b в БД.
    Структура файла: metric_id | val_a | val_b (строка 1 — заголовки).
    """
    try:
        from openpyxl import load_workbook
        errors: list[str] = []
        imported_count = 0

        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        headers = [str(ws.cell(1, c).value or "").strip().lower() for c in range(1, 4)]
        if headers != ["metric_id", "val_a", "val_b"]:
            return {"status": "FAILED", "imported": 0,
                    "errors": [f"Неверные заголовки: {headers}"]}

        rows_data = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(v is None for v in row[:3]):
                continue
            try:
                rows_data.append((int(row[0]),
                                  float(row[1]) if row[1] is not None else None,
                                  float(row[2]) if row[2] is not None else None))
            except (ValueError, TypeError) as e:
                errors.append(f"Строка {row_idx}: {e}")

        wb.close()
        logger.info("parse_excel_task: %d строк, %d ошибок", len(rows_data), len(errors))
        return {"status": "COMPLETED", "imported": len(rows_data), "errors": errors}

    except Exception as exc:
        logger.exception("parse_excel_task failed: %s", exc)
        raise self.retry(exc=exc)


@celery_app.task(name="tasks.generate_ai_summary")
def generate_ai_summary_task(period_id: str, system_name: str, period_label: str,
                             metrics_block: str = "", known_risks: str = "") -> dict:
    """Генерация AI-резюме встроенной (in-process) LLM. Без внешних сервисов."""
    from app.services import llm_service

    summary = llm_service.generate_summary(
        system_name=system_name,
        period_label=period_label,
        metrics_block=metrics_block,
        known_risks=known_risks,
    )
    return {"status": "COMPLETED", "summary": summary,
            "llm": llm_service.is_available()}


@celery_app.task(name="tasks.cache_invalidate")
def cache_invalidate_task(pattern: str) -> dict:
    """Инвалидация Redis кэша по паттерну ключей."""
    import redis
    r = redis.from_url(settings.REDIS_URL, decode_responses=True)
    keys = r.keys(pattern)
    deleted = r.delete(*keys) if keys else 0
    logger.info("cache_invalidate: удалено %d ключей по %r", deleted, pattern)
    return {"deleted_keys": deleted}
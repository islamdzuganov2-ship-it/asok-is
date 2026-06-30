import os
import uuid
from decimal import Decimal
from pathlib import Path
from typing import Any
from uuid import UUID

from celery.result import AsyncResult
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import require_role
from app.core.config import settings
from app.core.database import get_db
from app.models.assessment import AssessmentPeriod, AssessmentValue
from app.models.metric_catalog import MetricCatalog
from app.services.calculation_engine import calculate_metric, map_to_level
from app.services.excel_importer import import_matrices_from_workbook, import_metric_catalog_from_workbook
from app.workers.tasks import celery_app, parse_excel_task

router = APIRouter()

MAX_UPLOAD_SIZE = 10 * 1024 * 1024
# Сигнатура ZIP/OOXML — .xlsx это zip-контейнер. Проверяем содержимое, а не только расширение.
XLSX_MAGIC = b"PK\x03\x04"


def _validate_xlsx(filename: str | None, content: bytes) -> None:
    if not filename or not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")
    if len(content) > MAX_UPLOAD_SIZE:
        raise HTTPException(status_code=413, detail="File is larger than 10 MB")
    if not content.startswith(XLSX_MAGIC):
        raise HTTPException(status_code=400, detail="Файл не является корректным .xlsx (неверная сигнатура)")

HEADER_ALIASES = {
    "metric_id": {"metric_id", "id", "код", "код метрики", "ид метрики"},
    "characteristic": {"characteristic", "характеристика"},
    "subcharacteristic": {"subcharacteristic", "metric", "name", "метрика", "подхарактеристика", "название"},
    "val_a": {"val_a", "a", "значение a", "числитель", "факт"},
    "val_b": {"val_b", "b", "значение b", "знаменатель", "план", "база"},
    "expert_comment": {"expert_comment", "comment", "комментарий", "обоснование"},
}


def _normalize_header(value: object) -> str:
    return str(value or "").strip().lower().replace("\n", " ")


def _map_headers(row: tuple[object, ...]) -> dict[str, int]:
    headers: dict[str, int] = {}
    for index, cell in enumerate(row):
        normalized = _normalize_header(cell)
        for field, aliases in HEADER_ALIASES.items():
            if normalized in aliases and field not in headers:
                headers[field] = index
    return headers


def _cell(row: tuple[object, ...], headers: dict[str, int], key: str) -> object:
    index = headers.get(key)
    if index is None or index >= len(row):
        return None
    return row[index]


def _as_float(value: object) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return float(value)
    return float(str(value).replace(",", ".").strip())


async def _find_metric(
    db: AsyncSession,
    metric_id: object,
    characteristic: object,
    subcharacteristic: object,
) -> MetricCatalog | None:
    if metric_id not in (None, ""):
        try:
            metric = await db.get(MetricCatalog, int(metric_id))
            if metric is not None:
                return metric
        except (TypeError, ValueError):
            return None

    if subcharacteristic not in (None, ""):
        stmt = select(MetricCatalog).where(
            MetricCatalog.subcharacteristic.ilike(str(subcharacteristic).strip())
        )
        if characteristic not in (None, ""):
            stmt = stmt.where(MetricCatalog.characteristic.ilike(str(characteristic).strip()))
        result = await db.execute(stmt)
        return result.scalar_one_or_none()
    return None


@router.post("/upload")
async def upload_excel(
    period_id: str = Form(...),
    file: UploadFile = File(...),
    _: dict = Depends(require_role("TEST_ANALYST", "QUALITY_MANAGER", "ADMIN")),
) -> dict[str, str]:
    content = await file.read()
    _validate_xlsx(file.filename, content)

    os.makedirs(settings.UPLOAD_DIR, exist_ok=True)
    safe_name = f"{uuid.uuid4()}_{os.path.basename(file.filename)}"
    file_path = os.path.join(settings.UPLOAD_DIR, safe_name)
    with open(file_path, "wb") as target:
        target.write(content)

    task = parse_excel_task.delay(file_path, period_id)
    return {"task_id": task.id}


@router.get("/tasks/{task_id}")
async def get_task_status(task_id: str) -> dict[str, object]:
    result = AsyncResult(task_id, app=celery_app)
    payload: dict[str, object] = {"task_id": task_id, "status": result.status}
    if result.ready():
        if result.failed():
            payload["error"] = str(result.result)
        else:
            payload["result"] = result.result
    return payload


@router.post("/import-assessment")
async def import_assessment_excel(
    period_id: str = Form(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _: dict = Depends(require_role("TEST_ANALYST", "QUALITY_MANAGER", "ADMIN")),
) -> dict[str, Any]:
    content = await file.read()
    _validate_xlsx(file.filename, content)

    try:
        period_uuid = UUID(period_id)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail="Invalid period_id") from exc

    period = await db.get(AssessmentPeriod, period_uuid)
    if period is None:
        raise HTTPException(status_code=404, detail="Assessment period not found")

    os.makedirs(settings.UPLOAD_DIR, exist_ok=True)
    safe_name = f"{uuid.uuid4()}_{os.path.basename(file.filename)}"
    file_path = os.path.join(settings.UPLOAD_DIR, safe_name)
    with open(file_path, "wb") as target:
        target.write(content)

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    imported = 0
    skipped = 0
    errors: list[str] = []
    sheets: list[dict[str, Any]] = []

    try:
        for worksheet in workbook.worksheets:
            rows = worksheet.iter_rows(values_only=True)
            sheet_title = worksheet.title.strip().upper()
            
            # Маршрутизация парсинга в зависимости от целевой вкладки Excel-файла
            if "РИСК" in sheet_title:
                # Логика обработки и сохранения рисков в БД
                pass           
            header_row = next(rows, None)
            if header_row is None:
                sheets.append({"name": worksheet.title, "imported": 0, "skipped": 0})
                continue

            headers = _map_headers(header_row)
            required = {"val_a", "val_b"}
            if "metric_id" not in headers and "subcharacteristic" not in headers:
                errors.append(f"{worksheet.title}: не найдены колонки metric_id или subcharacteristic/name")
                continue
            if not required.issubset(headers):
                errors.append(f"{worksheet.title}: не найдены обязательные колонки val_a и val_b")
                continue

            sheet_imported = 0
            sheet_skipped = 0
            for row_index, row in enumerate(rows, start=2):
                if all(cell in (None, "") for cell in row):
                    continue
                try:
                    metric = await _find_metric(
                        db,
                        _cell(row, headers, "metric_id"),
                        _cell(row, headers, "characteristic"),
                        _cell(row, headers, "subcharacteristic"),
                    )
                    if metric is None:
                        sheet_skipped += 1
                        skipped += 1
                        errors.append(f"{worksheet.title}:{row_index}: метрика не найдена")
                        continue

                    result = await db.execute(
                        select(AssessmentValue).where(
                            AssessmentValue.period_id == period_uuid,
                            AssessmentValue.metric_id == metric.id,
                        )
                    )
                    value = result.scalar_one_or_none()
                    if value is None:
                        value = AssessmentValue(period_id=period_uuid, metric_id=metric.id, data_source="EXCEL")
                        db.add(value)

                    value.val_a = _as_float(_cell(row, headers, "val_a"))
                    value.val_b = _as_float(_cell(row, headers, "val_b"))
                    value.expert_comment = str(_cell(row, headers, "expert_comment") or "").strip() or None
                    formula_type = metric.formula_type.value if hasattr(metric.formula_type, "value") else str(metric.formula_type)
                    if value.val_a is not None and value.val_b is not None:
                        value.calculated_x = calculate_metric(float(value.val_a), float(value.val_b), formula_type)
                        value.quality_level = map_to_level(float(value.calculated_x))
                    value.data_source = "EXCEL"
                    sheet_imported += 1
                    imported += 1
                except Exception as exc:
                    sheet_skipped += 1
                    skipped += 1
                    errors.append(f"{worksheet.title}:{row_index}: {exc}")

            sheets.append({"name": worksheet.title, "imported": sheet_imported, "skipped": sheet_skipped})
        period.status = "CALCULATED"
        await db.commit()
    finally:
        workbook.close()

    return {
        "filename": file.filename,
        "period_id": period_id,
        "imported": imported,
        "skipped": skipped,
        "errors": errors[:100],
        "sheets": sheets,
    }


@router.post("/import-workbook")
async def import_workbook(
    period_id: str = Form(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _: dict = Depends(require_role("TEST_ANALYST", "QUALITY_MANAGER", "ADMIN")),
) -> dict[str, Any]:
    content = await file.read()
    _validate_xlsx(file.filename, content)

    try:
        period_uuid = UUID(period_id)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail="Invalid period_id") from exc

    period = await db.get(AssessmentPeriod, period_uuid)
    if period is None:
        raise HTTPException(status_code=404, detail="Assessment period not found")

    os.makedirs(settings.UPLOAD_DIR, exist_ok=True)
    safe_name = f"{uuid.uuid4()}_{os.path.basename(file.filename)}"
    file_path = os.path.join(settings.UPLOAD_DIR, safe_name)
    with open(file_path, "wb") as target:
        target.write(content)

    metrics_summary = await import_metric_catalog_from_workbook(db, Path(file_path))
    matrices_summary = await import_matrices_from_workbook(db, Path(file_path), period_uuid)
    await db.commit()
    return {
        "filename": file.filename,
        "period_id": period_id,
        "metrics": metrics_summary.as_dict(),
        "matrices": matrices_summary.as_dict(),
    }

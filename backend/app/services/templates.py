"""
Service for loading and parsing Excel template files.
"""
import os
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def get_templates_dir() -> Path:
    """Get the directory where templates are stored."""
    return Path(__file__).parent.parent.parent.parent


def load_excel_template(filename: str) -> list[dict[str, Any]]:
    """
    Load an Excel template file and return rows as list of dicts.
    
    Args:
        filename: Name of the Excel file in the project root
        
    Returns:
        List of row dictionaries
    """
    file_path = get_templates_dir() / filename
    
    if not file_path.exists():
        return []
    
    try:
        workbook = load_workbook(str(file_path), read_only=True, data_only=True)
        rows_data = []
        
        for worksheet in workbook.worksheets:
            sheet_rows = list(worksheet.iter_rows(values_only=True))
            if not sheet_rows:
                continue
            
            # Use first row as headers
            headers = [str(h or "").strip() for h in sheet_rows[0]]
            headers = [h for h in headers if h]  # Filter empty headers
            
            # Convert remaining rows to dicts
            for row in sheet_rows[1:]:
                if all(cell in (None, "") for cell in row):
                    continue
                row_dict = {}
                for idx, header in enumerate(headers):
                    if idx < len(row):
                        value = row[idx]
                        row_dict[header] = value if value is not None else ""
                rows_data.append(row_dict)
        
        workbook.close()
        return rows_data
    except Exception as e:
        print(f"Error loading template {filename}: {e}")
        return []


def load_metrics_template() -> list[dict[str, Any]]:
    """Load the metrics template."""
    return load_excel_template("шаблон заполнения метрик.xlsx")


def load_risks_template() -> list[dict[str, Any]]:
    """Load the risks template."""
    # Try different possible filenames
    for filename in [
        "шаблон предоставления информации для тест-менеджера после анализа LLM.xlsx",
        "таблица для заполнения v8.1.xlsx",
    ]:
        data = load_excel_template(filename)
        if data:
            return data
    return []


def load_quality_report_template() -> list[dict[str, Any]]:
    """Load the quality report template."""
    return load_excel_template("шаблон формирования детального отчета по каждой метрике.xlsx")


def load_system_quality_template() -> list[dict[str, Any]]:
    """Load the system quality template."""
    return load_excel_template("Качество системы в разрезе времени по всем характеристикам.xlsx")

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.assessment import AssessmentPeriod, AssessmentValue
from app.models.matrices import DefectMatrix, QualityPlanMatrix, RiskMatrix
from app.models.metric_catalog import FormulaType, MetricCatalog
from app.models.system import CriticalityClass, LifecycleStatus, System
from app.services.calculation_engine import map_to_level


DEFAULT_EXCEL_SYSTEM_CODE = "EXCEL_QUALITY"
DEFAULT_EXCEL_SYSTEM_NAME = "Система из Excel-файлов"


@dataclass
class ImportSummary:
    metrics: int = 0
    periods: int = 0
    values: int = 0
    risks: int = 0
    defects: int = 0
    plans: int = 0
    errors: list[str] | None = None

    def as_dict(self) -> dict[str, Any]:
        return {
            "metrics": self.metrics,
            "periods": self.periods,
            "values": self.values,
            "risks": self.risks,
            "defects": self.defects,
            "plans": self.plans,
            "errors": self.errors or [],
        }


def _text(value: Any) -> str:
    return str(value or "").replace("\n", " ").strip()


def _norm(value: Any) -> str:
    return " ".join(_text(value).lower().replace("ё", "е").split())


def _float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, str):
        value = value.replace("%", "").replace(",", ".").strip()
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number / 100 if number > 1 and "%" in str(value) else number


def _formula_type(value: Any) -> FormulaType:
    text = _norm(value)
    return FormulaType.INVERSE if "1 -" in text or "1-" in text else FormulaType.DIRECT


def _period_label(value: Any) -> str:
    text = _text(value).replace("кв", "Q").replace("КВ", "Q")
    parts = text.split()
    if len(parts) >= 3 and parts[1].upper() == "Q":
        return f"Q{parts[0]}-{parts[2]}"
    if len(parts) >= 2 and parts[1].lower().startswith("к"):
        return f"Q{parts[0]}-{parts[-1]}"
    return text


async def get_or_create_default_system(db: AsyncSession) -> System:
    result = await db.execute(select(System).where(System.code == DEFAULT_EXCEL_SYSTEM_CODE))
    system = result.scalar_one_or_none()
    if system is not None:
        return system

    system = System(
        name=DEFAULT_EXCEL_SYSTEM_NAME,
        code=DEFAULT_EXCEL_SYSTEM_CODE,
        status_lc=LifecycleStatus.OE,
        criticality_class=CriticalityClass.BUSINESS_OPERATIONAL,
        owner="Excel import",
        is_active=True,
    )
    db.add(system)
    await db.flush()
    return system


async def get_or_create_period(db: AsyncSession, system_id: UUID, period: str) -> tuple[AssessmentPeriod, bool]:
    result = await db.execute(
        select(AssessmentPeriod).where(
            AssessmentPeriod.system_id == system_id,
            AssessmentPeriod.period == period,
        )
    )
    existing = result.scalar_one_or_none()
    if existing is not None:
        return existing, False

    created = AssessmentPeriod(system_id=system_id, period=period, status="DRAFT")
    db.add(created)
    await db.flush()
    return created, True


async def get_or_create_metric(
    db: AsyncSession,
    characteristic: str,
    subcharacteristic: str,
    formula_type: FormulaType = FormulaType.DIRECT,
    description: str | None = None,
    data_source: str | None = None,
    metric_id: int | None = None,
) -> tuple[MetricCatalog, bool]:
    if metric_id is not None:
        existing = await db.get(MetricCatalog, metric_id)
        if existing is not None:
            existing.characteristic = characteristic or existing.characteristic
            existing.subcharacteristic = subcharacteristic or existing.subcharacteristic
            existing.formula_type = formula_type
            existing.description = description or existing.description
            existing.data_source = data_source or existing.data_source
            existing.is_active = True
            return existing, False

    result = await db.execute(
        select(MetricCatalog).where(
            MetricCatalog.characteristic == characteristic,
            MetricCatalog.subcharacteristic == subcharacteristic,
        )
    )
    existing = result.scalar_one_or_none()
    if existing is not None:
        existing.formula_type = formula_type
        existing.description = description or existing.description
        existing.data_source = data_source or existing.data_source
        existing.is_active = True
        return existing, False

    metric = MetricCatalog(
        id=metric_id,
        characteristic=characteristic,
        subcharacteristic=subcharacteristic,
        formula_type=formula_type,
        description=description,
        data_source=data_source,
        is_active=True,
    )
    db.add(metric)
    await db.flush()
    return metric, True


async def ensure_period_values(db: AsyncSession, period: AssessmentPeriod) -> int:
    metrics = list((await db.execute(select(MetricCatalog).where(MetricCatalog.is_active.is_(True)))).scalars().all())
    created = 0
    for metric in metrics:
        result = await db.execute(
            select(AssessmentValue).where(
                AssessmentValue.period_id == period.id,
                AssessmentValue.metric_id == metric.id,
            )
        )
        if result.scalar_one_or_none() is None:
            db.add(AssessmentValue(period_id=period.id, metric_id=metric.id, data_source="MANUAL"))
            created += 1
    return created


async def import_metric_catalog_from_workbook(db: AsyncSession, path: Path) -> ImportSummary:
    summary = ImportSummary(errors=[])
    workbook = load_workbook(path, data_only=True, read_only=False)
    try:
        sheet = next((ws for ws in workbook.worksheets if "АТТЕСТАЦИОННЫЙ" in ws.title.upper()), None)
        if sheet is None:
            return summary

        last_characteristic = ""
        for row in sheet.iter_rows(min_row=6, values_only=True):
            metric_number = row[2] if len(row) > 2 else None
            try:
                metric_id = int(metric_number)
            except (TypeError, ValueError):
                continue

            characteristic = _text(row[0]) or last_characteristic
            subcharacteristic = _text(row[1])
            if not characteristic or not subcharacteristic:
                continue
            last_characteristic = characteristic

            _, created = await get_or_create_metric(
                db,
                characteristic=characteristic,
                subcharacteristic=subcharacteristic,
                metric_id=metric_id,
                formula_type=_formula_type(row[4] if len(row) > 4 else ""),
                description=_text(row[3] if len(row) > 3 else "") or None,
                data_source=_text(row[7] if len(row) > 7 else "") or "Excel template",
            )
            summary.metrics += 1 if created else 0
    finally:
        workbook.close()
    return summary


async def import_quality_timeline(db: AsyncSession, path: Path, system: System | None = None) -> ImportSummary:
    summary = ImportSummary(errors=[])
    system = system or await get_or_create_default_system(db)
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook.worksheets[0]
        period_columns: list[tuple[int, str]] = []
        for col in range(4, sheet.max_column + 1):
            label = _period_label(sheet.cell(row=2, column=col).value)
            if label:
                period_columns.append((col, label))

        periods: dict[str, AssessmentPeriod] = {}
        for _, label in period_columns:
            period, created = await get_or_create_period(db, system.id, label)
            period.status = "CALCULATED"
            periods[label] = period
            summary.periods += 1 if created else 0

        current_characteristic = ""
        for row in range(4, sheet.max_row + 1):
            characteristic = _text(sheet.cell(row=row, column=2).value) or current_characteristic
            subcharacteristic = _text(sheet.cell(row=row, column=3).value)
            if not subcharacteristic or "целевой уровень" in _norm(characteristic):
                continue
            current_characteristic = characteristic
            metric, created = await get_or_create_metric(
                db,
                characteristic=characteristic,
                subcharacteristic=subcharacteristic,
                formula_type=FormulaType.DIRECT,
                data_source=path.name,
            )
            summary.metrics += 1 if created else 0

            for col, label in period_columns:
                score = _float(sheet.cell(row=row, column=col).value)
                if score is None:
                    continue
                period = periods[label]
                result = await db.execute(
                    select(AssessmentValue).where(
                        AssessmentValue.period_id == period.id,
                        AssessmentValue.metric_id == metric.id,
                    )
                )
                value = result.scalar_one_or_none()
                if value is None:
                    value = AssessmentValue(period_id=period.id, metric_id=metric.id)
                    db.add(value)
                value.val_a = round(score * 100, 2)
                value.val_b = 100
                value.calculated_x = round(max(0.0, min(1.0, score)), 4)
                value.quality_level = map_to_level(float(value.calculated_x))
                value.data_source = "EXCEL_TIMELINE"
                summary.values += 1
    finally:
        workbook.close()
    return summary


async def import_matrices_from_workbook(db: AsyncSession, path: Path, period_id: UUID) -> ImportSummary:
    summary = ImportSummary(errors=[])
    workbook = load_workbook(path, data_only=True, read_only=False)
    try:
        await db.execute(delete(RiskMatrix).where(RiskMatrix.period_id == period_id))
        await db.execute(delete(DefectMatrix).where(DefectMatrix.period_id == period_id))
        await db.execute(delete(QualityPlanMatrix).where(QualityPlanMatrix.period_id == period_id))

        for sheet in workbook.worksheets:
            title = sheet.title.upper()
            if "РИСК" in title:
                for row in sheet.iter_rows(min_row=7, values_only=True):
                    if not _text(row[0] if len(row) > 0 else "") or not _text(row[1] if len(row) > 1 else ""):
                        continue
                    risk = RiskMatrix(
                        period_id=period_id,
                        characteristic=_text(row[0]),
                        subcharacteristic=_text(row[1]),
                        risk_description=_text(row[2] if len(row) > 2 else "") or "-",
                        risk_consequence=_text(row[3] if len(row) > 3 else "") or "-",
                        mitigation_measures=_text(row[4] if len(row) > 4 else "") or "-",
                    )
                    db.add(risk)
                    summary.risks += 1
            elif "НЕДОСТАТ" in title:
                for row in sheet.iter_rows(min_row=18, values_only=True):
                    characteristic = _text(row[1] if len(row) > 1 else "") or _text(row[0] if len(row) > 0 else "")
                    defect = _text(row[4] if len(row) > 4 else "") or _text(row[3] if len(row) > 3 else "")
                    if not characteristic or not defect:
                        continue
                    db.add(
                        DefectMatrix(
                            period_id=period_id,
                            characteristic=characteristic,
                            digital_metric=_text(row[2] if len(row) > 2 else "") or None,
                            quality_metric_level=_text(row[3] if len(row) > 3 else "") or None,
                            defect_description=defect,
                        )
                    )
                    summary.defects += 1
            elif "ПЛАН" in title and "КАЧЕСТВ" in title:
                for row in sheet.iter_rows(min_row=16, values_only=True):
                    characteristic = _text(row[0] if len(row) > 0 else "")
                    task = _text(row[2] if len(row) > 2 else "")
                    if not characteristic or not task:
                        continue
                    db.add(
                        QualityPlanMatrix(
                            period_id=period_id,
                            characteristic=characteristic,
                            subcharacteristic=_text(row[1] if len(row) > 1 else "") or "-",
                            task_description=task,
                            internal_document=_text(row[3] if len(row) > 3 else "") or None,
                            assignee_fio=_text(row[4] if len(row) > 4 else "") or None,
                            assignee_role=_text(row[5] if len(row) > 5 else "") or None,
                            assignee_department=_text(row[6] if len(row) > 6 else "") or None,
                            deadline=_text(row[7] if len(row) > 7 else "") or "-",
                            profile_executor=_text(row[8] if len(row) > 8 else "") or None,
                            tech_debt_link=_text(row[9] if len(row) > 9 else "") or None,
                        )
                    )
                    summary.plans += 1
    finally:
        workbook.close()
    return summary


async def seed_project_excel_files(db: AsyncSession, project_root: Path) -> dict[str, Any]:
    candidates = [project_root, project_root / "seed_data", Path("/app/seed_data"), Path("/app")]
    def find_file(name: str) -> Path:
        for base in candidates:
            candidate = base / name
            if candidate.exists():
                return candidate
        return candidates[0] / name

    files = {
        "metrics_template": find_file("таблица для заполнения v8.1.xlsx"),
        "quality_timeline": find_file("Качество системы в разрезе времени по всем характеристикам.xlsx"),
    }
    result: dict[str, Any] = {}
    if files["metrics_template"].exists():
        summary = await import_metric_catalog_from_workbook(db, files["metrics_template"])
        result["metrics_template"] = summary.as_dict()
    if files["quality_timeline"].exists():
        summary = await import_quality_timeline(db, files["quality_timeline"])
        result["quality_timeline"] = summary.as_dict()
    await db.commit()
    return result

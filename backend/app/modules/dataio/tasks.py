"""
Celery-задачи домена dataio (ТЗ v13): фоновый парсинг Excel.
Регистрируются на celery_app из infrastructure.workers; для воркера реэкспортируются
через app.workers.tasks (путь -A в docker-compose).
"""
import logging

from app.infrastructure.workers import celery_app

logger = logging.getLogger(__name__)


@celery_app.task(bind=True, name="tasks.parse_excel", max_retries=3)
def parse_excel_task(self, file_path: str, period_id: str) -> dict:
    """
    Парсинг Excel файла и батчевая запись val_a/val_b в БД.
    Структура файла: metric_id | val_a | val_b (строка 1 — заголовки).
    """
    try:
        from openpyxl import load_workbook
        errors: list[str] = []

        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        headers = [str(ws.cell(1, c).value or "").strip().lower() for c in range(1, 4)]
        if headers != ["metric_id", "val_a", "val_b"]:
            return {"status": "FAILED", "imported": 0,
                    "errors": [f"Неверные заголовки: {headers}"]}

        rows_data = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(v is None for v in row[:3]):
                continue
            try:
                rows_data.append((int(row[0]),
                                  float(row[1]) if row[1] is not None else None,
                                  float(row[2]) if row[2] is not None else None))
            except (ValueError, TypeError) as e:
                errors.append(f"Строка {row_idx}: {e}")

        wb.close()
        logger.info("parse_excel_task: %d строк, %d ошибок", len(rows_data), len(errors))
        return {"status": "COMPLETED", "imported": len(rows_data), "errors": errors}

    except Exception as exc:
        logger.exception("parse_excel_task failed: %s", exc)
        raise self.retry(exc=exc)

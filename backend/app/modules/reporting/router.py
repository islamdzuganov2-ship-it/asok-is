"""
REST API домена reporting (ТЗ v13): управленческий дашборд, LLM-аналитика по мерам,
Excel-матрицы периода, статус LLM.

Транзитная зависимость: llm_service ещё в app.services (переезжает задачей #15).
"""
import asyncio
import os
from datetime import datetime
from collections import defaultdict
from io import BytesIO
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from pydantic import BaseModel, ConfigDict
from pydantic.alias_generators import to_camel
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.infrastructure.database import get_db
from app.modules.assessment.models import AssessmentPeriod, AssessmentValue
from app.modules.iam import get_current_user, require_permission
from app.modules.llm import brain as llm_brain
from app.modules.quality import (
    CHARACTERISTICS,
    DEFAULT_CRITICALITY_WEIGHTS,
    QUALITY_MODEL,
    SUBCHAR_WEIGHTS,
    SubcharScore,
    canonical_characteristic,
    portfolio_score,
    weighted_system_score,
)
from app.modules.reporting.models import DefectMatrix, QualityPlanMatrix, RiskMatrix
from app.modules.reporting.schemas import (
    DashboardDataOut,
    DefectMatrixRow,
    FullExcelMatricesOut,
    PeriodsUsedOut,
    ProblematicSystemOut,
    QualityPlanMatrixRow,
    RiskMatrixRow,
)
from app.modules.governance import list_proposals  # кросс-доменный фасад (T-15: метки мер)
from app.modules.llm import decisions as llm_decisions
from app.modules.llm import gate as llm_gate
from app.modules.llm import personas as llm_personas
from app.modules.llm import pipeline as llm_pipeline
from app.modules.llm import reasoning as llm_reasoning
from app.modules.llm import selfcheck as llm_selfcheck
from app.modules.llm import service as llm_service
from app.modules.llm.tasks import llm_selfcheck_task
from app.modules.risk import RiskBase, risks_for_characteristics
from app.modules.systems import System
from app.shared.periods import period_sort_key

router = APIRouter()


@router.get("/llm-status")
async def get_llm_status(_: dict = Depends(get_current_user)) -> dict:
    """Статус встроенной LLM: паспорт загруженной модели (архитектура/контекст) + мозг."""
    return llm_service.model_info()


@router.get("/llm-models")
async def get_llm_models(_: dict = Depends(get_current_user)) -> dict:
    """Список GGUF-моделей в каталоге (для переключения/диагностики): имя, размер, выбранная."""
    return {"models": llm_service.list_models()}


@router.get("/llm-pipeline")
async def get_llm_pipeline(_: dict = Depends(get_current_user)) -> dict:
    """Матрица конвейера RAG и обучения (ТЗ v18 п.1): источники контекста, механизмы, уровни.

    Открыта всем аутентифицированным: это описание архитектуры, а не данные оценки. Отчёт
    самооценки (llm-quality) — напротив, только для суперадминистратора.
    """
    return {
        **llm_pipeline.summary(),
        "personas": llm_personas.catalog(),
        "decision_matrices": llm_decisions.matrices(),
    }


@router.get("/llm-quality")
async def get_llm_quality_report(
    _: dict = Depends(require_permission("view.admin.llm_quality")),
) -> dict:
    """Последний отчёт самооценки LLM по ISO/IEC 25010 (ТЗ v18 п.10). Только суперадминистратор."""
    report = llm_selfcheck.latest()
    return {
        "report": report,
        "history": llm_selfcheck.history(),
        "schedule": llm_selfcheck.schedule_description(),
    }


@router.post("/llm-quality/run")
async def run_llm_quality_check(
    mode: str = "full",
    _: dict = Depends(require_permission("llm.quality.run")),
) -> dict:
    """Ручной запуск самооценки (ТЗ v18 п.10). Только суперадминистратор.

    mode="full" уходит в фоновую задачу: полный прогон с инференсом занимает минуты и
    блокировал бы HTTP-запрос (llama.cpp сериализует инференс). mode="static" выполняется
    сразу — интроспективные пробы занимают доли секунды.
    """
    if mode not in ("full", "static"):
        raise HTTPException(status_code=400, detail="mode должен быть full или static")
    if mode == "static":
        return {"status": "COMPLETED", "mode": mode, "report": llm_selfcheck.run(
            mode="static", trigger="manual")}
    task = llm_selfcheck_task.delay(mode="full", trigger="manual")
    return {"status": "QUEUED", "mode": mode, "task_id": task.id,
            "hint": "полный прогон выполняется в фоне; результат появится в /reports/llm-quality"}


@router.post("/llm-reload")
async def reload_llm(_: dict = Depends(require_permission("llm.reload"))) -> dict:
    """Горячая перезагрузка модели без рестарта контейнера (после подмены файла). Только ADMIN."""
    return llm_service.reload()


class ConclusionFeedbackIn(BaseModel):
    """Обратная связь человека по заключению LLM (обучение «резервного мозга»)."""
    fingerprint: str
    verdict: str                      # accept | reject | edit
    edited_text: str | None = None    # для verdict=edit — исправленный экспертом текст
    user: str | None = None


@router.post("/conclusion-feedback")
async def conclusion_feedback(payload: ConclusionFeedbackIn,
                              _: dict = Depends(require_permission("assessment.review"))) -> dict:
    """Принимает оценку/правку заключения → пишет в «резервный мозг» (влияет на будущий recall)."""
    try:
        entry = llm_brain.record_feedback(
            payload.fingerprint, payload.verdict, payload.edited_text or "", payload.user or "",
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    return {"status": "ok", "feedback": entry, "brain": llm_brain.stats()}


class MeasuresAnalyticsItem(BaseModel):
    characteristic: str
    count: int
    systems: int | None = None
    avg_score: float | None = None


class MeasureCardIn(BaseModel):
    """Карточка меры из governance-контура (MeasureDecisionModal) — первичный факт для конвейера."""
    system: str
    characteristic: str
    title: str
    rationale: str | None = None      # обоснование (профсуждение в карточке)
    expectation: str | None = None    # что ожидается от ЛПР
    owner: str | None = None
    due: str | None = None
    score: float | None = None


class MeasuresReasoningIn(BaseModel):
    """Новый формат запроса: агрегаты + сами карточки мер (первичные данные для анализа)."""
    items: list[MeasuresAnalyticsItem] = []
    cards: list[MeasureCardIn] = []


def _measures_blocks(items: list[MeasuresAnalyticsItem], cards: list[MeasureCardIn]) -> str:
    """Блок фактов о мерах: сводка по характеристикам + карточки мер (обоснование/ответственный/срок)."""
    lines = [
        f"{i.characteristic} | мер: {i.count}"
        + (f", ИС: {i.systems}" if i.systems else "")
        + (f", ср.балл: {round(i.avg_score)}%" if i.avg_score is not None else "")
        for i in sorted(items, key=lambda x: x.count, reverse=True)
    ]
    for c in cards[:12]:
        detail = f"{c.system} | {c.characteristic} | {c.title}"
        if c.rationale:
            detail += f": {c.rationale}"
        extras = [x for x in (
            f"балл {round(c.score)}%" if c.score is not None and c.score >= 0 else None,
            f"ответственный {c.owner}" if c.owner else None,
            f"срок {c.due}" if c.due else None,
        ) if x]
        if extras:
            detail += f" ({', '.join(extras)})"
        lines.append(detail)
    return "\n".join(lines)


@router.post("/measures-analytics")
async def measures_analytics(
    payload: MeasuresReasoningIn | list[MeasuresAnalyticsItem],
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permission("view.reports")),
) -> dict:
    """Аналитика по МЕРАМ через конвейер многоаспектного аналитического рассуждения (BL-005).

    Принимает новый формат {items, cards} (карточки мер — первичный источник) и легаси-массив
    агрегатов (обратная совместимость). Ответ сохраняет прежние поля (analytics/llm/mapped_risks)
    и дополняется аудируемой трассой (reasoning) и уверенностью (confidence).

    ТЗ v18: вывод адресуется персоне по роли запросившего (personas.py).
    """
    items = payload if isinstance(payload, list) else payload.items
    cards = [] if isinstance(payload, list) else payload.cards
    if not items and not cards:
        return {"analytics": "Активных мер нет — аналитика не сформирована.",
                "llm": llm_service.is_available(), "mapped_risks": [],
                "reasoning": None, "confidence": "низкая"}

    measures_block = _measures_blocks(items, cards)
    chars = sorted({i.characteristic for i in items} | {c.characteristic for c in cards})
    risk_rows = await risks_for_characteristics(db, chars, limit=8)
    risks_block = "\n".join(f"- {r.title}: {r.mitigation or '—'}" for r in risk_rows)

    # Rule Engine: самая просевшая характеристика по среднему баллу мер → триггер Severity.
    scores = [i.avg_score for i in items if i.avg_score is not None]
    q_proxy = (min(scores) / 100.0) if scores else None
    gate_result = llm_gate.evaluate_gate(q=q_proxy)
    rules_block = gate_result.as_block()

    # Персона адресата — из роли запросившего (RBAC → персона, см. modules/llm/personas.py).
    roles = current_user.get("roles") or []
    persona = llm_personas.resolve(roles[0] if roles else "")

    result = await asyncio.to_thread(
        llm_reasoning.generate_reasoned_conclusion,
        "ИТ-ландшафт банка", "текущий период",
        "",             # профсуждения в этом потоке не передаются (их поток — judgment-conclusion)
        risks_block, "", measures_block, "", rules_block,
        # Поток сводный (по всему ландшафту), поэтому класс критичности отдельной ИС здесь
        # не применим и в матрицу триажа не передаётся — она отработает по мягкому классу.
        gate_result.severity, "", 0, 0, persona.code,
    )
    return {
        "analytics": result["conclusion"],
        "llm": llm_service.is_available(),
        "mapped_risks": [{"title": r.title, "characteristic": r.characteristic} for r in risk_rows],
        "reasoning": result["trace"],
        "confidence": result["confidence"],
        "fingerprint": result["fingerprint"],
        "fired_rules": [ln.lstrip("- ") for ln in rules_block.splitlines() if ln.strip()],
        "persona": {"code": persona.code, "title": persona.title, "audience": persona.audience},
        "decisions": result["decisions"],
    }


@router.get("/assessment-period/{period_id}/matrices", response_model=FullExcelMatricesOut)
async def get_period_excel_matrices(
    period_id: UUID,
    db: AsyncSession = Depends(get_db),
    _: dict = Depends(get_current_user),
) -> FullExcelMatricesOut:
    period = await db.get(AssessmentPeriod, period_id)
    if period is None:
        raise HTTPException(status_code=404, detail="Assessment period not found")

    risks = list(
        (
            await db.execute(
                select(RiskMatrix).where(RiskMatrix.period_id == period_id).order_by(RiskMatrix.id)
            )
        )
        .scalars()
        .all()
    )
    defects = list(
        (
            await db.execute(
                select(DefectMatrix).where(DefectMatrix.period_id == period_id).order_by(DefectMatrix.id)
            )
        )
        .scalars()
        .all()
    )
    plans = list(
        (
            await db.execute(
                select(QualityPlanMatrix).where(QualityPlanMatrix.period_id == period_id).order_by(QualityPlanMatrix.id)
            )
        )
        .scalars()
        .all()
    )

    return FullExcelMatricesOut(
        period_id=period_id,
        risks=[
            RiskMatrixRow(
                characteristic=row.characteristic,
                subcharacteristic=row.subcharacteristic,
                risk_description=row.risk_description,
                risk_consequence=row.risk_consequence,
                mitigation_measures=row.mitigation_measures,
            )
            for row in risks
        ],
        defects=[
            DefectMatrixRow(
                id=row.id,
                characteristic=row.characteristic,
                digital_metric=row.digital_metric,
                quality_metric_level=row.quality_metric_level,
                defect_description=row.defect_description,
            )
            for row in defects
        ],
        plan=[
            QualityPlanMatrixRow(
                id=row.id,
                characteristic=row.characteristic,
                subcharacteristic=row.subcharacteristic,
                task_description=row.task_description,
                internal_document=row.internal_document,
                assignee_fio=row.assignee_fio,
                assignee_role=row.assignee_role,
                assignee_department=row.assignee_department,
                deadline=row.deadline,
                profile_executor=row.profile_executor,
                tech_debt_link=row.tech_debt_link,
            )
            for row in plans
        ],
    )


def _score_to_bucket(value: float | None) -> int:
    if value is None:
        return 0
    if value >= 0.81:
        return 5
    if value >= 0.61:
        return 4
    if value >= 0.41:
        return 3
    if value >= 0.21:
        return 2
    if value > 0:
        return 1
    return 0


@router.get("/executive-dashboard", response_model=DashboardDataOut)
async def get_executive_dashboard(db: AsyncSession = Depends(get_db),
                                  _: dict = Depends(get_current_user)) -> DashboardDataOut:
    result = await db.execute(
        select(AssessmentValue)
        .options(
            selectinload(AssessmentValue.metric),
            selectinload(AssessmentValue.period).selectinload(AssessmentPeriod.system),
        )
        .join(AssessmentPeriod, AssessmentValue.period_id == AssessmentPeriod.id)
        .join(System, AssessmentPeriod.system_id == System.id)
        .where(System.is_deleted.is_(False), System.is_active.is_(True))
    )
    all_values = list(result.scalars().all())

    # Сквозная консистентность (DEF-12): управленческий дашборд считает ТОЛЬКО последний период
    # каждой ИС — как /assessments/dashboard. Хронология периодов — по (год, квартал) (DEF-13).
    latest_period_per_system: dict[UUID, UUID] = {}
    for value in sorted(all_values, key=lambda v: period_sort_key(v.period.period), reverse=True):
        latest_period_per_system.setdefault(value.period.system_id, value.period_id)
    values = [v for v in all_values if latest_period_per_system.get(v.period.system_id) == v.period_id]

    measured = [float(value.calculated_x) for value in values if value.calculated_x is not None]
    if not measured:
        return DashboardDataOut(
            globalHealthScore=0.0,
            aiInsights="Нет данных для расчета управленческой панели.",
            heatmapData=[],
            xAxisLabels=[],
            yAxisLabels=[],
            problematicSystems=[],
            periodsUsed=PeriodsUsedOut(distinct=[], earliest=None, latest=None, bySystem={}),
        )

    system_names = sorted({value.period.system.name for value in values})

    # ТЗ v19 УК-07 (Р-6): та же двухуровневая свёртка, что и /assessments/dashboard — раньше
    # здесь было НЕЗАВИСИМОЕ плоское среднее по ВСЕМ строкам разом (DEF-12 обещал консистентность
    # «как в /assessments/dashboard» комментарием, а не общим кодом — веса подхарактеристик и
    # критичности до этого экрана не доходили; ExecutiveDashboard.tsx ходит именно сюда).
    cells: dict[tuple[str, str], list[float]] = defaultdict(list)
    low_counts: dict[UUID, int] = defaultdict(int)
    systems_by_id: dict[UUID, System] = {}
    present_chars: set[str] = set()
    subchar_x: dict[str, dict[tuple[str, str], float]] = defaultdict(dict)
    period_label_by_system: dict[str, str] = {}
    for value in values:
        systems_by_id[value.period.system.id] = value.period.system
        period_label_by_system[value.period.system.name] = value.period.period
        if value.calculated_x is None:
            continue
        # Нормализация имени характеристики к модели 25010 (DEF-02): теплокарта = 8 характеристик, как в моках.
        canon = canonical_characteristic(value.metric.characteristic)
        if canon is None:
            continue
        score = float(value.calculated_x)
        present_chars.add(canon)
        cells[(value.period.system.name, canon)].append(score)
        subchar_x[value.period.system.name][(canon, value.metric.subcharacteristic)] = score * 100
        if score < 0.41:
            low_counts[value.period.system.id] += 1

    crit_by_name: dict[str, str] = {
        s.name: (s.criticality_class.value if hasattr(s.criticality_class, "value") else str(s.criticality_class))
        for s in systems_by_id.values()
    }
    system_scores: dict[str, float | None] = {}
    for name in system_names:
        subchar_scores = [
            SubcharScore(
                characteristic=char_title, subcharacteristic=sub,
                weight=SUBCHAR_WEIGHTS[(char_title, sub)],
                x=subchar_x.get(name, {}).get((char_title, sub)),
            )
            for char_title, subs_def in QUALITY_MODEL
            for sub, _formula in subs_def
        ]
        system_scores[name] = weighted_system_score(subchar_scores).score
    portfolio = portfolio_score(system_scores, crit_by_name, DEFAULT_CRITICALITY_WEIGHTS)
    global_score = round(portfolio.score, 2) if portfolio.score is not None else 0.0

    # Канонические характеристики в фиксированном порядке модели (только присутствующие).
    characteristic_names = [c for c in CHARACTERISTICS if c in present_chars]
    system_index = {name: index for index, name in enumerate(system_names)}
    characteristic_index = {name: index for index, name in enumerate(characteristic_names)}

    heatmap = [
        [
            characteristic_index[characteristic],
            system_index[system_name],
            _score_to_bucket(sum(scores) / len(scores)),
        ]
        for (system_name, characteristic), scores in cells.items()
    ]

    distinct_periods = sorted(set(period_label_by_system.values()), key=period_sort_key)
    periods_used = PeriodsUsedOut(
        distinct=distinct_periods,
        earliest=distinct_periods[0] if distinct_periods else None,
        latest=distinct_periods[-1] if distinct_periods else None,
        bySystem=period_label_by_system,
    )

    problematic = [
        ProblematicSystemOut(
            id=system.id,
            name=system.name,
            criticality=system.criticality_class.value
            if hasattr(system.criticality_class, "value")
            else str(system.criticality_class),
            lowMetricsCount=count,
        )
        for system_id, count in sorted(low_counts.items(), key=lambda item: item[1], reverse=True)
        if (system := systems_by_id.get(system_id)) is not None
    ][:10]

    # --- Управленческое резюме: встроенная LLM с обоснованием из базы рисков ---
    char_avg: dict[str, list[float]] = defaultdict(list)
    for (system_name, characteristic), scores in cells.items():
        char_avg[characteristic].append(sum(scores) / len(scores))
    char_pct = {c: round(sum(v) / len(v) * 100) for c, v in char_avg.items() if v}
    metrics_block = "\n".join(
        f"{c} | средняя по ИС | {pct}%" for c, pct in sorted(char_pct.items(), key=lambda kv: kv[1])
    )
    weak_chars = [c for c, pct in sorted(char_pct.items(), key=lambda kv: kv[1])[:2]]

    known_risks = ""
    if weak_chars:
        risk_rows = await risks_for_characteristics(db, weak_chars, limit=5)
        known_risks = "\n".join(
            f"- {r.title}: {r.mitigation or '—'}" for r in risk_rows
        )

    ai_insights = await asyncio.to_thread(
        llm_service.generate_summary,
        "ИТ-ландшафт банка",
        "текущий период",
        metrics_block or f"Средний интегральный показатель — {global_score}%.",
        known_risks,
    )

    return DashboardDataOut(
        globalHealthScore=global_score,
        aiInsights=ai_insights,
        heatmapData=heatmap,
        xAxisLabels=characteristic_names,
        yAxisLabels=system_names,
        problematicSystems=problematic,
        periodsUsed=periods_used,
    )


class _CamelOut(BaseModel):
    """Базовая Out-схема с camelCase-алиасами (как DashboardDataOut — для консистентности фронта)."""
    model_config = ConfigDict(alias_generator=to_camel, populate_by_name=True)


class DynamicsPoint(_CamelOut):
    period: str
    integral: float                       # интегральный показатель качества за период, %
    characteristics: dict[str, float]     # характеристика → средний %


class MeasureMarker(_CamelOut):
    """Метка меры на шкале динамики (T-15): когда по характеристике поставлена мера."""
    characteristic: str
    created_at: str
    title: str
    status: str


class MeasureEffect(_CamelOut):
    """Фактическая результативность меры: ΔScore «до/после» (ДЕФ-32, БТ-134, T-15).

    До сих пор в системе жил только ПЛАНОВЫЙ `expected_delta_score`, который вводит человек,
    а фактического сравнения не было — на графике рисовались лишь метки мер, и ответить
    «привела ли мера к улучшению» приходилось глазами.
    """
    characteristic: str
    title: str
    status: str
    period_before: str
    period_after: str
    score_before: float
    score_after: float
    delta: float                 # процентные пункты: положительный — улучшение
    verdict: str                 # «улучшение» | «без изменений» | «ухудшение»


class SystemDynamicsOut(_CamelOut):
    system_id: UUID
    system_name: str
    points: list[DynamicsPoint]
    measures: list[MeasureMarker]  # метки мер по характеристикам (T-15: наложить на ряд)
    effects: list[MeasureEffect] = []  # ДЕФ-32: посчитанный ΔScore по каждой мере



# Порог, ниже которого изменение считаем шумом округления, а не результатом меры.
EFFECT_NOISE_PP = 1.0


def _measure_effects(points: list[DynamicsPoint], measures: list[MeasureMarker]) -> list[MeasureEffect]:
    """ΔScore по каждой мере: балл характеристики ДО принятия меры и в следующем периоде.

    «До» — последний период, начавшийся не позже даты меры; «после» — первый период строго
    следом. Если одной из точек нет (мера свежая или период ещё не закрыт), меру пропускаем:
    честнее не показать результат, чем показать ноль и выдать его за «без изменений».
    """
    if len(points) < 2:
        return []
    order = {p.period: i for i, p in enumerate(points)}
    effects: list[MeasureEffect] = []
    for m in measures:
        if not m.created_at:
            continue
        # Период, в котором мера была принята: ищем по хронологии периодов оценки.
        idx_before = None
        for i, p in enumerate(points):
            if p.period in order and p.characteristics.get(m.characteristic) is not None:
                # Берём последнюю точку, которая по хронологии НЕ позже даты меры.
                if _period_starts_before(p.period, m.created_at):
                    idx_before = i
        if idx_before is None or idx_before + 1 >= len(points):
            continue
        before, after = points[idx_before], points[idx_before + 1]
        score_before = before.characteristics.get(m.characteristic)
        score_after = after.characteristics.get(m.characteristic)
        if score_before is None or score_after is None:
            continue
        delta = round(score_after - score_before, 1)
        verdict = ("улучшение" if delta > EFFECT_NOISE_PP
                   else "ухудшение" if delta < -EFFECT_NOISE_PP
                   else "без изменений")
        effects.append(MeasureEffect(
            characteristic=m.characteristic, title=m.title, status=m.status,
            period_before=before.period, period_after=after.period,
            score_before=score_before, score_after=score_after,
            delta=delta, verdict=verdict,
        ))
    return effects


def _period_starts_before(period: str, iso_datetime: str) -> bool:
    """Начался ли квартал `period` (вида «Q3-2026») не позже даты в ISO-формате."""
    try:
        year, quarter = period_sort_key(period)
    except Exception:  # noqa: BLE001 — период произвольного вида: сравнить не можем
        return False
    try:
        stamp = datetime.fromisoformat(iso_datetime.replace("Z", "+00:00"))
    except ValueError:
        return False
    measure_quarter = (stamp.year, (stamp.month - 1) // 3 + 1)
    return (year, quarter) <= measure_quarter

@router.get("/system-dynamics", response_model=SystemDynamicsOut)
async def system_dynamics(system_id: UUID, db: AsyncSession = Depends(get_db),
                          _: dict = Depends(get_current_user)) -> SystemDynamicsOut:
    """Динамика качества ИС по периодам (T-15 — эффективность мер + live-режим «Динамики качества»).

    Интегральный показатель и средние по 8 характеристикам (ISO 25010) за каждый квартал в
    хронологии (period_sort_key). Позволяет оценить ЭФФЕКТИВНОСТЬ принятой меры: стало ли лучше по
    её характеристике в периодах после внедрения (мера накладывается на ряд по characteristic+дате).
    """
    system = await db.get(System, system_id)
    if system is None:
        raise HTTPException(status_code=404, detail="Система не найдена")
    result = await db.execute(
        select(AssessmentValue)
        .options(selectinload(AssessmentValue.metric), selectinload(AssessmentValue.period))
        .join(AssessmentPeriod, AssessmentValue.period_id == AssessmentPeriod.id)
        .where(AssessmentPeriod.system_id == system_id)
    )
    by_period: dict[str, dict[str, list[float]]] = defaultdict(lambda: defaultdict(list))
    for v in result.scalars().all():
        if v.calculated_x is None:
            continue
        canon = canonical_characteristic(v.metric.characteristic)
        if canon:
            by_period[v.period.period][canon].append(float(v.calculated_x))

    points: list[DynamicsPoint] = []
    for period in sorted(by_period, key=period_sort_key):
        chars = {c: round(sum(s) / len(s) * 100, 1) for c, s in by_period[period].items()}
        integral = round(sum(chars.values()) / len(chars), 1) if chars else 0.0
        points.append(DynamicsPoint(period=period, integral=integral, characteristics=chars))

    # Метки мер (T-15): когда и по какой характеристике поставлена мера — чтобы на фронте
    # наложить на ряд и увидеть, изменилось ли качество по характеристике ПОСЛЕ внедрения.
    proposals = await list_proposals(db, system=system.name)
    measures = [
        MeasureMarker(
            characteristic=canonical_characteristic(p.characteristic) or (p.characteristic or "—"),
            created_at=p.created_at.isoformat() if p.created_at else "",
            title=p.risk_title or p.characteristic or "Мера",
            status=p.status,
        )
        for p in proposals if p.characteristic
    ]
    effects = _measure_effects(points, measures)
    return SystemDynamicsOut(system_id=system_id, system_name=system.name, points=points,
                             measures=measures, effects=effects)


_XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _write_sheet(ws, headers: list[str], rows: list[list]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)



_PDF_MEDIA = "application/pdf"


def _pdf_font() -> str:
    """Шрифт с кириллицей. Встроенные Helvetica/Times её не содержат — текст вышел бы «квадратами».

    Берём DejaVuSans из reportlab, он поставляется вместе с библиотекой.
    """
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.rl_config import TTFSearchPath

    name = "DejaVuSans"
    if name in pdfmetrics.getRegisteredFontNames():
        return name
    for base in list(TTFSearchPath) + ["/usr/share/fonts/truetype/dejavu"]:
        candidate = os.path.join(str(base), "DejaVuSans.ttf")
        if os.path.isfile(candidate):
            pdfmetrics.registerFont(TTFont(name, candidate))
            return name
    return "Helvetica"   # крайний случай: латиница выведется, кириллица — нет


def _build_summary_pdf(period_label: str, system_name: str, blocks: list[tuple[str, list[str], list[list]]]) -> BytesIO:
    """Сводный отчёт одним PDF: заголовок + таблицы по блокам (ДЕФ-18, T-14, БТ-283)."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    font = _pdf_font()
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=12 * mm, rightMargin=12 * mm, topMargin=12 * mm, bottomMargin=12 * mm,
        title=f"АСОК ИС — сводный отчёт {period_label}",
    )
    base = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=base["Title"], fontName=font, fontSize=15, spaceAfter=4)
    h2 = ParagraphStyle("h2", parent=base["Heading2"], fontName=font, fontSize=11, spaceBefore=8, spaceAfter=4)
    small = ParagraphStyle("small", parent=base["Normal"], fontName=font, fontSize=8, leading=10)

    story: list = [
        Paragraph(f"АСОК ИС — сводный отчёт по качеству", h1),
        Paragraph(f"ИС «{system_name}» · период {period_label}", small),
        Spacer(1, 6),
    ]
    for title, headers, rows in blocks:
        story.append(Paragraph(f"{title} — строк: {len(rows)}", h2))
        if not rows:
            story.append(Paragraph("Данных нет.", small))
            continue
        data = [[Paragraph(str(h), small) for h in headers]]
        for row in rows:
            data.append([Paragraph("" if c is None else str(c), small) for c in row])
        table = Table(data, repeatRows=1, hAlign="LEFT")
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EAEBEE")),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#C9CDD3")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(table)
    doc.build(story)
    buf.seek(0)
    return buf


async def _load_period_matrices(db: AsyncSession, period_id: UUID):
    """Значения оценки и три реестра периода одним местом.

    Выгрузки xlsx и PDF читают одно и то же; раньше четыре одинаковых запроса были
    скопированы в оба эндпоинта. Общий загрузчик убирает дубль и держит роутер в пределах
    бюджета обращений к ORM (tests/test_router_purity.py).
    """
    vals = list((await db.execute(
        select(AssessmentValue).options(selectinload(AssessmentValue.metric))
        .where(AssessmentValue.period_id == period_id)
    )).scalars().all())
    risks = list((await db.execute(
        select(RiskMatrix).where(RiskMatrix.period_id == period_id).order_by(RiskMatrix.id)
    )).scalars().all())
    defects = list((await db.execute(
        select(DefectMatrix).where(DefectMatrix.period_id == period_id).order_by(DefectMatrix.id)
    )).scalars().all())
    plans = list((await db.execute(
        select(QualityPlanMatrix).where(QualityPlanMatrix.period_id == period_id).order_by(QualityPlanMatrix.id)
    )).scalars().all())
    return vals, risks, defects, plans

@router.get("/export/{period_id}/xlsx")
async def export_period_xlsx(period_id: UUID, db: AsyncSession = Depends(get_db),
                             _: dict = Depends(require_permission("view.reports"))) -> StreamingResponse:
    """Выгрузка реестров периода в .xlsx (T-14): характеристики качества, риски, недостатки, план.

    Управленческая выгрузка «одним файлом» (4 листа) — то, что в ТЗ v11 R2.1 просилось экспортом
    per-реестр. Данные — те же таблицы, что показывает вкладка «Отчёты и реестры».
    """
    period = await db.get(AssessmentPeriod, period_id)
    if period is None:
        raise HTTPException(status_code=404, detail="Период оценки не найден")

    vals, risks, defects, plans = await _load_period_matrices(db, period_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Характеристики качества"
    _write_sheet(
        ws,
        ["Характеристика", "Подхарактеристика", "A", "B", "X", "Уровень", "Комментарий"],
        [[v.metric.characteristic, v.metric.subcharacteristic, v.val_a, v.val_b,
          float(v.calculated_x) if v.calculated_x is not None else None,
          v.quality_level, v.expert_comment] for v in vals],
    )
    _write_sheet(
        wb.create_sheet("Риски"),
        ["Характеристика", "Подхарактеристика", "Описание риска", "Последствие", "Меры минимизации"],
        [[r.characteristic, r.subcharacteristic, r.risk_description, r.risk_consequence, r.mitigation_measures] for r in risks],
    )
    _write_sheet(
        wb.create_sheet("Недостатки"),
        ["N", "Характеристика", "Цифровой показатель", "Уровень качества", "Описание недостатка"],
        [[r.id, r.characteristic, r.digital_metric, r.quality_metric_level, r.defect_description] for r in defects],
    )
    _write_sheet(
        wb.create_sheet("План качества"),
        ["N", "Характеристика", "Подхарактеристика", "Задача", "ВНД", "Ответственный", "Срок"],
        [[r.id, r.characteristic, r.subcharacteristic, r.task_description, r.internal_document, r.assignee_fio, r.deadline] for r in plans],
    )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"asok_report_{period.period}.xlsx".replace(" ", "_")
    return StreamingResponse(
        buf, media_type=_XLSX_MEDIA,
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@router.get("/export/{period_id}/pdf")
async def export_period_pdf(period_id: UUID, db: AsyncSession = Depends(get_db),
                            _: dict = Depends(require_permission("view.reports"))) -> StreamingResponse:
    """Сводный отчёт периода одним PDF (ДЕФ-18, T-14, БТ-283).

    Экспорт в xlsx работал, PDF — нет, хотя в требованиях он стоял рядом («экспорт xlsx +
    сводный PDF»). Содержимое то же, что в xlsx-выгрузке: характеристики качества, риски,
    недостатки и план качества — но в виде, пригодном для рассылки и печати.
    """
    period = await db.get(AssessmentPeriod, period_id)
    if period is None:
        raise HTTPException(status_code=404, detail="Период оценки не найден")
    system = await db.get(System, period.system_id)

    vals, risks, defects, plans = await _load_period_matrices(db, period_id)

    blocks = [
        ("Характеристики качества",
         ["Характеристика", "Подхарактеристика", "A", "B", "X", "Уровень"],
         [[v.metric.characteristic, v.metric.subcharacteristic, v.val_a, v.val_b,
           float(v.calculated_x) if v.calculated_x is not None else None, v.quality_level]
          for v in vals]),
        ("Риски",
         ["Характеристика", "Подхарактеристика", "Описание риска", "Последствие", "Меры минимизации"],
         [[r.characteristic, r.subcharacteristic, r.risk_description, r.risk_consequence,
           r.mitigation_measures] for r in risks]),
        ("Недостатки",
         ["N", "Характеристика", "Показатель", "Уровень", "Описание"],
         [[r.id, r.characteristic, r.digital_metric, r.quality_metric_level, r.defect_description]
          for r in defects]),
        ("План качества",
         ["N", "Характеристика", "Задача", "Ответственный", "Срок"],
         [[r.id, r.characteristic, r.task_description, r.assignee_fio, r.deadline] for r in plans]),
    ]
    # Сборка PDF синхронная и не быстрая на больших периодах — уводим с event loop (ДЕФ-26).
    buf = await asyncio.to_thread(
        _build_summary_pdf, period.period, system.name if system else "—", blocks,
    )
    fname = f"asok_report_{period.period}.pdf".replace(" ", "_")
    return StreamingResponse(
        buf, media_type=_PDF_MEDIA,
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )

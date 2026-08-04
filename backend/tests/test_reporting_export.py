"""Тест выгрузки реестров периода в .xlsx (T-14, ТЗ v11 R2.1).

Проверяет эндпоинт `GET /reports/export/{period_id}/xlsx` в режиме in-process
(роутер-функция + тестовая сессия, как в test_assessment_correction.py):
  • файл собирается и открывается openpyxl-ом (валидный .xlsx);
  • четыре листа в фиксированном порядке — характеристики/риски/недостатки/план;
  • данные каждого реестра попадают на свой лист (шапка + строки);
  • несуществующий период → 404.
"""
import uuid
from io import BytesIO

import pytest
from fastapi import HTTPException
from openpyxl import load_workbook

from app.modules.assessment.models import AssessmentPeriod, AssessmentValue
from app.modules.quality import QUALITY_PAIRS, FormulaType, MetricCatalog, calculate_metric, map_to_level
from app.modules.reporting.models import DefectMatrix, QualityPlanMatrix, RiskMatrix
from app.modules.reporting.router import export_period_xlsx
from app.modules.systems import CriticalityClass, System

SHEETS = ["Характеристики качества", "Риски", "Недостатки", "План качества"]


async def _system(db, name="АБС Core") -> System:
    system = System(
        id=uuid.uuid4(), name=name, code=f"S-{uuid.uuid4().hex[:8]}",
        criticality_class=CriticalityClass.BUSINESS_CRITICAL,
    )
    db.add(system)
    await db.flush()
    return system


async def _metrics(db) -> list[MetricCatalog]:
    rows = [
        MetricCatalog(characteristic=c, subcharacteristic=s, formula_type=FormulaType(f), is_active=True)
        for c, s, f in QUALITY_PAIRS
    ]
    db.add_all(rows)
    await db.flush()
    return rows


async def _period(db, system, metrics, quarter="Q2-2026", x=0.5) -> AssessmentPeriod:
    period = AssessmentPeriod(id=uuid.uuid4(), system_id=system.id, period=quarter, status="CALCULATED")
    db.add(period)
    await db.flush()
    for metric in metrics:
        formula = metric.formula_type.value
        b = 100
        a = round(b * x) if formula == "DIRECT" else round(b * (1 - x))
        real_x = calculate_metric(a, b, formula)
        db.add(AssessmentValue(
            id=uuid.uuid4(), period_id=period.id, metric_id=metric.id,
            val_a=a, val_b=b, calculated_x=real_x, quality_level=map_to_level(real_x),
            data_source="TEST",
        ))
    await db.flush()
    return period


async def _read_xlsx(response) -> "load_workbook":
    """Собрать тело StreamingResponse и открыть как книгу openpyxl."""
    chunks = [chunk async for chunk in response.body_iterator]
    data = b"".join(c if isinstance(c, bytes) else c.encode() for c in chunks)
    return load_workbook(BytesIO(data))


async def test_export_xlsx_has_four_sheets_with_data(db_session):
    system = await _system(db_session)
    metrics = await _metrics(db_session)
    period = await _period(db_session, system, metrics)

    db_session.add(RiskMatrix(
        period_id=period.id, characteristic="Надёжность", subcharacteristic="Безотказность",
        risk_description="Рост отказов после релиза", risk_consequence="Простой бизнес-процесса",
        mitigation_measures="Программа стабилизации, сокращение MTTR",
    ))
    db_session.add(DefectMatrix(
        period_id=period.id, characteristic="Производительность",
        digital_metric="p95 latency", quality_metric_level="Низкий",
        defect_description="Деградация времени отклика под нагрузкой",
    ))
    db_session.add(QualityPlanMatrix(
        period_id=period.id, characteristic="Сопровождаемость", subcharacteristic="Тестируемость",
        task_description="Поднять покрытие интеграционными тестами", internal_document="ВНД-42",
        assignee_fio="Иванов И.И.", deadline="30.09.2026",
    ))
    await db_session.flush()

    response = await export_period_xlsx(period.id, db_session)
    assert response.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert f"asok_report_{period.period}.xlsx" in response.headers["Content-Disposition"]

    wb = await _read_xlsx(response)
    # Четыре листа в фиксированном порядке (управленческая выгрузка «одним файлом»).
    assert wb.sheetnames == SHEETS

    quality = wb["Характеристики качества"]
    assert quality.max_row == len(metrics) + 1                      # шапка + строка на каждую пару
    assert quality.cell(row=1, column=1).value == "Характеристика"  # шапка на месте

    risks = wb["Риски"]
    assert risks.max_row == 2
    assert risks.cell(row=2, column=1).value == "Надёжность"
    assert "MTTR" in risks.cell(row=2, column=5).value

    defects = wb["Недостатки"]
    assert defects.cell(row=2, column=2).value == "Производительность"

    plan = wb["План качества"]
    assert plan.cell(row=2, column=6).value == "Иванов И.И."
    assert plan.cell(row=2, column=7).value == "30.09.2026"


async def test_export_xlsx_empty_registers_still_valid(db_session):
    # Пустые реестры (риски/недостатки/план не заведены) — лист есть, только шапка.
    system = await _system(db_session)
    metrics = await _metrics(db_session)
    period = await _period(db_session, system, metrics)

    wb = await _read_xlsx(await export_period_xlsx(period.id, db_session))
    assert wb.sheetnames == SHEETS
    assert wb["Риски"].max_row == 1          # только шапка
    assert wb["Недостатки"].max_row == 1
    assert wb["План качества"].max_row == 1


async def test_export_xlsx_unknown_period_404(db_session):
    with pytest.raises(HTTPException) as err:
        await export_period_xlsx(uuid.uuid4(), db_session)
    assert err.value.status_code == 404
